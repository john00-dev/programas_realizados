'''

Este script permite capturar URLs de un video en reproducción, detener el video, copiar 
la URL al portapapeles y guardarla en un archivo Excel.

'''
import os
import pyperclip
import pyautogui
import openpyxl
import time

# Espera 10 segundos antes de ejecutar el código (puedes ajustar este tiempo según sea necesario)
time.sleep(10)

# Nombre del archivo Excel
excel_file = 'urls.xlsx'

# Verificar si el archivo Excel existe
if os.path.exists(excel_file):
    # Si el archivo existe, cargarlo y agregar la URL
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
else:
    # Si el archivo no existe, crear uno nuevo y añadir encabezados
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["URL"])  # Añadir encabezado


# Lista para mantener un registro de las URLs y sus conteos
url_registros = []

# Número máximo de registros permitidos para una URL
max_registros = 10

# Bucle para capturar múltiples URLs
while True:
    try:
        # Parar el video
        x, y = 226, 218
        pyautogui.moveTo(x, y)
        pyautogui.click()
        pyautogui.press('down')

        # Mover la barra del buscador
        x, y = 303, 81
        pyautogui.moveTo(x, y,0.1)
        pyautogui.click()

        # Copiar la URL
        pyautogui.hotkey('ctrl', 'c')
        url = pyperclip.paste()

        # Verificar si la URL ya está en el registro y no ha alcanzado el límite de repeticiones
        url_repetida = next((registro for registro in url_registros if registro["url"] == url), None)
        if url_repetida:
            if url_repetida["conteo"] >= 8:
                print(f"La URL {url} se ha repetido 5 veces. Saliendo del bucle.")
                break
            else:
                url_repetida["conteo"] += 1
        else:
            # Si la URL no está en el registro, agregarla con un conteo de 1
            url_registros.append({"url": url, "conteo": 1})

        # Verificar si la URL ya ha alcanzado el límite de registros
        if len(url_registros) > max_registros:
            url_registros.pop(0)  # Eliminar el primer registro más antiguo

        # Añadir la URL a la siguiente fila disponible en la columna A
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=url)

        # Guardar cambios en el archivo Excel periódicamente para evitar la pérdida de datos en caso de un fallo
        if next_row % 100 == 0:
            workbook.save(excel_file)

    
    except Exception as e:
        # Manejar cualquier error que ocurra durante la automatización
        print(f"Error: {e}")
        break

# Guardar los cambios finales en el archivo Excel
workbook.save(excel_file)